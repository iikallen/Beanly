import csv
import hashlib
import re
import zipfile
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from beanly.core.money import MAX_NUMERIC_20_6_MINOR
from beanly.modules.onboarding.application.dto import CanonicalImportDraft, CanonicalImportEntity
from beanly.modules.onboarding.domain.enums import (
    ImportEntityType,
    ImportSourceType,
    UploadSourceType,
)
from beanly.modules.onboarding.domain.exceptions import ImportFileTypeInvalid, ImportParseFailed

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 10_000
MAX_ZIP_ENTRIES = 500
MAX_ZIP_EXPANDED_BYTES = 50 * 1024 * 1024
_XLS_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def parse_spreadsheet(
    content: bytes,
    filename: str,
    content_type: str | None,
    requested_source: UploadSourceType,
    mapping: dict[str, str] | None = None,
) -> tuple[CanonicalImportDraft, ImportSourceType, str]:
    if not content or len(content) > MAX_UPLOAD_BYTES:
        raise ImportFileTypeInvalid("Spreadsheet is empty or exceeds 10 MB")
    _validate_declared_type(content_type)
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        rows = {"Products": _csv_rows(content)}
        detected = ImportSourceType.GENERIC_SPREADSHEET
    elif suffix == ".xlsx":
        if not content.startswith(b"PK"):
            raise ImportFileTypeInvalid("XLSX magic does not match its extension")
        _validate_xlsx_archive(content)
        rows = _xlsx_rows(content)
        detected = _detect_source(rows, requested_source)
    elif suffix == ".xls" and requested_source is UploadSourceType.POSTER:
        if not content.startswith(_XLS_MAGIC):
            raise ImportFileTypeInvalid("XLS magic does not match its extension")
        rows = _xls_rows(content)
        detected = ImportSourceType.POSTER_EXPORT
    else:
        raise ImportFileTypeInvalid("Only CSV/XLSX, or Poster XLS, is supported")
    rows = _remap_sheets(rows, mapping or {})
    if detected is ImportSourceType.GENERIC_SPREADSHEET and not _sheet(rows, "Products"):
        first_name = next(iter(rows), None)
        rows = {"Products": rows[first_name]} if first_name else rows
    if detected is ImportSourceType.GENERIC_SPREADSHEET:
        rows = _generic_profile(rows)
    draft = _poster(rows) if detected is ImportSourceType.POSTER_EXPORT else _beanly(rows)
    return draft, detected, hashlib.sha256(content).hexdigest()


def inspect_spreadsheet(
    content: bytes,
    filename: str,
    content_type: str | None,
    requested_source: UploadSourceType,
) -> dict[str, object]:
    if not content or len(content) > MAX_UPLOAD_BYTES:
        raise ImportFileTypeInvalid("Spreadsheet is empty or exceeds 10 MB")
    _validate_declared_type(content_type)
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        rows = {"Products": _csv_rows(content)}
        detected = ImportSourceType.GENERIC_SPREADSHEET
    elif suffix == ".xlsx":
        if not content.startswith(b"PK"):
            raise ImportFileTypeInvalid("XLSX magic does not match its extension")
        _validate_xlsx_archive(content)
        rows = _xlsx_rows(content)
        detected = _detect_source(rows, requested_source)
    elif suffix == ".xls" and requested_source is UploadSourceType.POSTER:
        if not content.startswith(_XLS_MAGIC):
            raise ImportFileTypeInvalid("XLS magic does not match its extension")
        rows = _xls_rows(content)
        detected = ImportSourceType.POSTER_EXPORT
    else:
        raise ImportFileTypeInvalid("Only CSV/XLSX, or Poster XLS, is supported")
    return {
        "file_hash": hashlib.sha256(content).hexdigest(),
        "source_type": detected,
        "sheets": [
            {"name": name, "columns": list(values[0]) if values else []}
            for name, values in rows.items()
        ],
        "mapping_required": detected is ImportSourceType.GENERIC_SPREADSHEET,
    }


def _validate_declared_type(content_type: str | None) -> None:
    if content_type and content_type.casefold() not in {
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        raise ImportFileTypeInvalid("Declared spreadsheet content type is not supported")


def _remap_sheets(
    sheets: dict[str, list[dict[str, object]]], mapping: dict[str, str]
) -> dict[str, list[dict[str, object]]]:
    if not mapping:
        return sheets
    if len(mapping) > 64 or any(
        not isinstance(source, str)
        or not isinstance(target, str)
        or len(source) > 150
        or len(target) > 100
        for source, target in mapping.items()
    ):
        raise ImportParseFailed("Column mapping is invalid or too large")
    normalized = {_header(source): _header(target) for source, target in mapping.items()}
    if len(set(normalized.values())) != len(normalized):
        raise ImportParseFailed("Multiple source columns cannot map to one canonical field")
    source_columns = {
        key for records in sheets.values() for record in records[:1] for key in record
    }
    allowed_targets = {
        "category",
        "product",
        "variant",
        "sku",
        "price",
        "location",
        "available",
        "description",
        "name",
        "unit",
        "opening quantity",
        "unit cost kzt",
    }
    if not set(normalized).issubset(source_columns) or not set(normalized.values()).issubset(
        allowed_targets
    ):
        raise ImportParseFailed("Column mapping references an unknown source or target field")
    remapped = {
        name: [
            {normalized.get(key, key): value for key, value in record.items()} for record in records
        ]
        for name, records in sheets.items()
    }
    columns = {key for records in remapped.values() for record in records[:1] for key in record}
    product_profile = {"category", "product", "price"}.issubset(columns)
    inventory_profile = {"name", "unit"}.issubset(columns)
    if not product_profile and not inventory_profile:
        raise ImportParseFailed(
            "Mapping must define category/product/price or inventory name/unit"
        )
    return remapped


def _generic_profile(
    sheets: dict[str, list[dict[str, object]]],
) -> dict[str, list[dict[str, object]]]:
    """Route a one-sheet generic file into the existing canonical parsers."""
    if _sheet(sheets, "Inventory"):
        return sheets
    products = _sheet(sheets, "Products")
    if products and {"name", "unit"}.issubset(products[0]) and not {
        "category",
        "product",
        "price",
    }.issubset(products[0]):
        return {"Inventory": products}
    return sheets


def _csv_rows(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise ImportParseFailed("CSV must be valid UTF-8") from exc
    reader = csv.reader(StringIO(text, newline=""))
    values = list(reader)
    if not values:
        raise ImportParseFailed("CSV is empty")
    return _records(values[0], values[1:])


def _xlsx_rows(content: bytes) -> dict[str, list[dict[str, object]]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as exc:
        raise ImportParseFailed("XLSX cannot be opened") from exc
    result: dict[str, list[dict[str, object]]] = {}
    total = 0
    for sheet in workbook.worksheets:
        raw_rows = list(sheet.iter_rows(values_only=False))
        if not raw_rows:
            continue
        for row in raw_rows:
            if any(cell.data_type == "f" for cell in row):
                raise ImportParseFailed("Formula cells are not accepted")
        values = [[cell.value for cell in row] for row in raw_rows]
        header_index = _first_header_row(values)
        records = _records(values[header_index], values[header_index + 1 :])
        total += len(records)
        if total > MAX_ROWS:
            raise ImportParseFailed("Spreadsheet exceeds 10000 data rows")
        result[sheet.title] = records
    if not result:
        raise ImportParseFailed("Workbook contains no data")
    return result


def _xls_rows(content: bytes) -> dict[str, list[dict[str, object]]]:
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise ImportParseFailed("XLS cannot be opened") from exc
    result: dict[str, list[dict[str, object]]] = {}
    total = 0
    for sheet in workbook.sheets():
        if sheet.nrows == 0:
            continue
        values = [sheet.row_values(row) for row in range(sheet.nrows)]
        header_index = _first_header_row(values)
        records = _records(values[header_index], values[header_index + 1 :])
        total += len(records)
        if total > MAX_ROWS:
            raise ImportParseFailed("Spreadsheet exceeds 10000 data rows")
        result[sheet.name] = records
    if not result:
        raise ImportParseFailed("Workbook contains no data")
    return result


def _first_header_row(rows: list[list[object]]) -> int:
    """Poster templates include a deliberately blank leading row."""
    for index, row in enumerate(rows[:20]):
        if sum(value not in (None, "") for value in row) >= 2:
            return index
    raise ImportParseFailed("Spreadsheet has no recognizable header row")


def _records(headers: list[object], rows: list[list[object]]) -> list[dict[str, object]]:
    normalized = [_header(value) for value in headers]
    if not any(normalized) or len(set(normalized)) != len(normalized):
        raise ImportParseFailed("Headers are empty or duplicated")
    result: list[dict[str, object]] = []
    for row in rows:
        values = list(row) + [None] * (len(normalized) - len(row))
        if not any(value not in (None, "") for value in values):
            continue
        record = dict(zip(normalized, values, strict=False))
        for value in record.values():
            if isinstance(value, str) and value.lstrip().startswith(_FORMULA_PREFIXES):
                raise ImportParseFailed("Formula-like text is not accepted")
        result.append(record)
    return result


def _beanly(sheets: dict[str, list[dict[str, object]]]) -> CanonicalImportDraft:
    products = _sheet(sheets, "Products")
    if not products and not _sheet(sheets, "Inventory"):
        raise ImportParseFailed("Products or Inventory sheet is required")
    entities: list[CanonicalImportEntity] = []
    categories: dict[str, str] = {}
    products_by_name: dict[str, str] = {}
    variants: dict[tuple[str, str], str] = {}
    order = 0
    for row in products:
        category = _text(row, "category")
        product = _text(row, "product")
        variant = _text(row, "variant", default="Default")
        category_key = categories.setdefault(_slug(category), f"category:{_slug(category)}")
        if not any(value.source_key == category_key for value in entities):
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.CATEGORY, category_key, {"name": category}, sort_order=order
                )
            )
            order += 1
        product_key = products_by_name.setdefault(_slug(product), f"product:{_slug(product)}")
        if not any(value.source_key == product_key for value in entities):
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.PRODUCT,
                    product_key,
                    {
                        "category_key": category_key,
                        "name": product,
                        "description": _optional(row.get("description")),
                        "status": "DRAFT",
                    },
                    sort_order=order,
                )
            )
            order += 1
        variant_key = f"variant:{_slug(product)}:{_slug(variant)}"
        variants[(product.casefold(), variant.casefold())] = variant_key
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.VARIANT,
                variant_key,
                {
                    "product_key": product_key,
                    "name": variant,
                    "sku": _optional(row.get("sku")),
                    "price_minor": _major_to_minor(row.get("price")),
                    "is_default": not any(
                        value.entity_type is ImportEntityType.VARIANT
                        and value.payload.get("product_key") == product_key
                        for value in entities
                    ),
                },
                sort_order=order,
            )
        )
        order += 1
        if row.get("location") not in (None, ""):
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.LOCATION_PRICE,
                    f"location-price:{variant_key}:{_slug(str(row['location']))}",
                    {
                        "variant_key": variant_key,
                        "location_name": str(row["location"]).strip(),
                        "available": _strict_bool(row.get("available"), default=True),
                        "price_minor": _major_to_minor(row.get("price")),
                    },
                    sort_order=order,
                )
            )
            order += 1
    inventory_keys: dict[str, str] = {}
    for row in _sheet(sheets, "Inventory"):
        name = _text(row, "name")
        key = f"inventory:{_slug(str(row.get('sku') or name))}"
        inventory_keys[name.casefold()] = key
        source_unit = _text(row, "unit")
        unit, quantity_factor = _base_unit(source_unit)
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.INVENTORY_ITEM,
                key,
                {"name": name, "sku": _optional(row.get("sku")), "base_unit": unit},
                sort_order=order,
            )
        )
        order += 1
        if row.get("opening quantity") not in (None, ""):
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.OPENING_BALANCE,
                    f"opening:{key}",
                    {
                        "inventory_item_key": key,
                        "quantity": _decimal_text(
                            Decimal(str(row.get("opening quantity"))) * quantity_factor
                        ),
                        "unit": unit,
                        "unit_cost_minor": _major_to_minor(row.get("unit cost kzt"))
                        if row.get("unit cost kzt") not in (None, "")
                        else None,
                        "unit_cost_base_factor": format(quantity_factor, "f"),
                    },
                    sort_order=order,
                )
            )
            order += 1
    recipe_groups: dict[str, list[dict[str, object]]] = {}
    for row in _sheet(sheets, "Recipes"):
        product = _text(row, "product")
        variant = _text(row, "variant", default="Default")
        variant_key = variants.get((product.casefold(), variant.casefold()))
        item_name = _text(row, "inventory item")
        item_key = inventory_keys.get(item_name.casefold())
        if not variant_key or not item_key:
            raise ImportParseFailed("Recipe references an unknown variant or inventory item")
        recipe_groups.setdefault(variant_key, []).append(
            {
                "inventory_item_key": item_key,
                "quantity": _decimal_text(row.get("quantity")),
                "unit": _text(row, "unit"),
            }
        )
    for variant_key, components in recipe_groups.items():
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.RECIPE,
                f"recipe:{variant_key}",
                {"variant_key": variant_key, "components": components, "review_required": True},
                warning_codes=("DRAFT_RECIPE_REVIEW_REQUIRED",),
                sort_order=order,
            )
        )
        order += 1
    group_keys: set[str] = set()
    for row in _sheet(sheets, "Modifiers"):
        product = _text(row, "product")
        variant = _text(row, "variant", default="Default")
        variant_key = variants.get((product.casefold(), variant.casefold()))
        if not variant_key:
            raise ImportParseFailed("Modifier references an unknown product variant")
        group_name = _text(row, "group")
        group_key = f"modifier-group:{variant_key}:{_slug(group_name)}"
        if group_key not in group_keys:
            group_keys.add(group_key)
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.MODIFIER_GROUP,
                    group_key,
                    {
                        "variant_key": variant_key,
                        "name": group_name,
                        "selection_type": _text(row, "selection type"),
                        "min_selections": 0,
                        "max_selections": (
                            1 if _text(row, "selection type") == "SINGLE" else 10
                        ),
                    },
                    sort_order=order,
                )
            )
            order += 1
        item_name = _text(row, "inventory item")
        item_key = inventory_keys.get(item_name.casefold())
        if not item_key:
            raise ImportParseFailed("Modifier references an unknown inventory item")
        option = _text(row, "option")
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.MODIFIER_OPTION,
                f"modifier-option:{group_key}:{_slug(option)}",
                {
                    "group_key": group_key,
                    "name": option,
                    "price_delta_minor": _major_to_minor(row.get("price delta")),
                    "inventory_deltas": [
                        {
                            "inventory_item_key": item_key,
                            "quantity": _decimal_text(row.get("quantity delta")),
                            "unit": _text(row, "unit"),
                        }
                    ],
                },
                sort_order=order,
            )
        )
        order += 1
    return CanonicalImportDraft("Beanly spreadsheet", 1, tuple(entities))


def _poster(sheets: dict[str, list[dict[str, object]]]) -> CanonicalImportDraft:
    entities: list[CanonicalImportEntity] = []
    categories: set[str] = set()
    products: dict[str, str] = {}
    variants: set[str] = set()
    inventory: dict[str, tuple[str, str]] = {}
    openings: set[str] = set()
    recipe_groups: dict[str, dict[str, dict[str, object]]] = {}
    order = 0

    def ensure_category(category: str) -> str:
        nonlocal order
        category_key = f"category:poster:{_slug(category)}"
        if category_key not in categories:
            categories.add(category_key)
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.CATEGORY, category_key, {"name": category}, sort_order=order
                )
            )
            order += 1
        return category_key

    def ensure_product(name: str, category: str, poster_id: str | None) -> str:
        nonlocal order
        identity = poster_id or name
        product_key = f"product:poster:{_slug(identity)}"
        if product_key not in products:
            category_key = ensure_category(category)
            products[product_key] = name
            entities.append(
                CanonicalImportEntity(
                    ImportEntityType.PRODUCT,
                    product_key,
                    {
                        "category_key": category_key,
                        "name": name,
                        "status": "DRAFT",
                        "source_external_id": poster_id,
                    },
                    warning_codes=("POSTER_REAL_FIXTURE_UNVERIFIED",),
                    sort_order=order,
                )
            )
            order += 1
        return product_key

    def ensure_variant(
        product_key: str,
        name: str,
        external_id: str,
        price: object,
    ) -> str:
        nonlocal order
        variant_key = f"variant:poster:{_slug(product_key)}:{_slug(external_id)}"
        if variant_key in variants:
            return variant_key
        variants.add(variant_key)
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.VARIANT,
                variant_key,
                {
                    "product_key": product_key,
                    "name": name,
                    "sku": None,
                    "price_minor": _major_to_minor(price),
                    "is_default": not any(
                        value.entity_type is ImportEntityType.VARIANT
                        and value.payload.get("product_key") == product_key
                        for value in entities
                    ),
                    "source_external_id": external_id,
                },
                sort_order=order,
            )
        )
        order += 1
        return variant_key

    def ensure_inventory(name: str, source_unit: str, sku: str | None = None) -> tuple[str, str]:
        nonlocal order
        base_unit, _ = _base_unit(source_unit)
        identity = (sku or name).casefold()
        existing = inventory.get(identity)
        if existing:
            if existing[1] != base_unit:
                raise ImportParseFailed("Poster ingredient uses incompatible units")
            return existing
        item_key = f"inventory:poster:{_slug(sku or name)}"
        inventory[identity] = (item_key, base_unit)
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.INVENTORY_ITEM,
                item_key,
                {"name": name, "sku": sku, "base_unit": base_unit},
                warning_codes=("POSTER_REAL_FIXTURE_UNVERIFIED",),
                sort_order=order,
            )
        )
        order += 1
        return item_key, base_unit

    for rows in sheets.values():
        for row in rows:
            dish_name = _poster_value(row, "блюдо", "название блюда", "product_name")
            ingredient_name = _poster_value(
                row, "состав", "ингредиент", "название ингредиента", "ingredient_name"
            )
            if dish_name and ingredient_name:
                category = _poster_value(row, "категория", "category") or "Imported"
                poster_id = _poster_value(
                    row,
                    "posterid product_id (не менять!)",
                    "posterid product_id",
                    "product_id",
                )
                product_key = ensure_product(dish_name, category, poster_id)
                variant_key = ensure_variant(
                    product_key,
                    dish_name,
                    f"base-{poster_id or dish_name}",
                    _poster_value(row, "цена", "price") or 0,
                )
                explicit_unit = _poster_value(row, "ед. изм.", "единица", "unit")
                quantity_value = _poster_value(
                    row, "брутто, г", "брутто", "нетто, г", "нетто", "quantity"
                )
                if quantity_value is None:
                    raise ImportParseFailed("Poster recipe component quantity is required")
                source_unit = explicit_unit or "g"
                item_key, base_unit = ensure_inventory(ingredient_name, source_unit)
                _, factor = _base_unit(source_unit)
                quantity = _poster_decimal(quantity_value) * factor
                components = recipe_groups.setdefault(variant_key, {})
                component = components.setdefault(
                    item_key,
                    {"inventory_item_key": item_key, "quantity": Decimal(0), "unit": base_unit},
                )
                component["quantity"] = Decimal(str(component["quantity"])) + quantity
                continue

            unit = _poster_value(row, "unit", "ед. изм.", "единица", "единица измерения")
            inventory_count = _poster_value(
                row, "inventory count", "остаток", "количество", "quantity"
            )
            generic_name = _poster_value(
                row, "name", "наименование", "название ингредиента", "ингредиент"
            )
            if generic_name and unit and inventory_count is not None:
                sku = _poster_value(row, "sku", "артикул", "ingredient_sku")
                item_key, base_unit = ensure_inventory(generic_name, unit, sku)
                if item_key not in openings:
                    openings.add(item_key)
                    _, factor = _base_unit(unit)
                    entities.append(
                        CanonicalImportEntity(
                            ImportEntityType.OPENING_BALANCE,
                            f"opening:{item_key}",
                            {
                                "inventory_item_key": item_key,
                                "quantity": _decimal_text(
                                    _poster_decimal(inventory_count) * factor
                                ),
                                "unit": base_unit,
                                "unit_cost_minor": _major_to_minor(
                                    _poster_value(row, "value", "cost_price", "себестоимость")
                                )
                                if _poster_value(
                                    row, "value", "cost_price", "себестоимость"
                                )
                                is not None
                                else None,
                                "unit_cost_base_factor": format(factor, "f"),
                            },
                            warning_codes=("POSTER_REAL_FIXTURE_UNVERIFIED",),
                            sort_order=order,
                        )
                    )
                    order += 1
                continue

            name = _poster_value(row, "name", "блюдо", "название", "название блюда")
            if not name:
                continue
            category = _poster_value(row, "category", "категория") or "Imported"
            poster_id = _poster_value(
                row,
                "posterid product_id (не менять!)",
                "posterid product_id",
                "product_id",
            )
            product_key = ensure_product(name, category, poster_id)
            modifier_id = _poster_value(
                row,
                "posterid modificator_id (не менять!)",
                "posterid modificator_id",
                "modificator_id",
            )
            external_id = modifier_id or f"base-{poster_id or name}"
            ensure_variant(
                product_key,
                name,
                external_id,
                _poster_value(row, "price", "цена", "цена продажи") or 0,
            )

    for variant_key, by_item in recipe_groups.items():
        components = [
            {
                **component,
                "quantity": _decimal_text(component["quantity"]),
            }
            for component in by_item.values()
        ]
        entities.append(
            CanonicalImportEntity(
                ImportEntityType.RECIPE,
                f"recipe:{variant_key}",
                {"variant_key": variant_key, "components": components, "review_required": True},
                warning_codes=(
                    "DRAFT_RECIPE_REVIEW_REQUIRED",
                    "POSTER_REAL_FIXTURE_UNVERIFIED",
                ),
                sort_order=order,
            )
        )
        order += 1
    if not entities:
        raise ImportParseFailed("Poster export has no recognized rows")
    return CanonicalImportDraft("Poster export", None, tuple(entities))


def _poster_value(row: dict[str, object], *aliases: str) -> str | None:
    for alias in aliases:
        value = row.get(_header(alias))
        if value not in (None, ""):
            return str(value).strip()
    return None


def _poster_decimal(value: object) -> Decimal:
    text = str(value).strip().casefold().replace(" ", "").replace(",", ".")
    match = re.match(r"^[+]?(\d+(?:\.\d+)?)", text)
    if not match:
        raise ImportParseFailed("Poster quantity must be a positive decimal")
    amount = Decimal(match.group(1))
    if not amount.is_finite() or amount <= 0:
        raise ImportParseFailed("Poster quantity must be a positive decimal")
    return amount


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise ImportFileTypeInvalid("XLSX contains too many archive entries")
            if sum(info.file_size for info in infos) > MAX_ZIP_EXPANDED_BYTES:
                raise ImportFileTypeInvalid("XLSX expanded content is too large")
            if any(".." in Path(info.filename).parts for info in infos):
                raise ImportFileTypeInvalid("XLSX contains an unsafe archive path")
    except zipfile.BadZipFile as exc:
        raise ImportFileTypeInvalid("XLSX archive is invalid") from exc


def _detect_source(
    sheets: dict[str, list[dict[str, object]]], requested: UploadSourceType
) -> ImportSourceType:
    if requested is UploadSourceType.POSTER:
        return ImportSourceType.POSTER_EXPORT
    if requested is UploadSourceType.BEANLY_SPREADSHEET:
        return ImportSourceType.BEANLY_SPREADSHEET
    product_rows = _sheet(sheets, "Products")
    if product_rows and {"category", "product", "price"}.issubset(product_rows[0]):
        return ImportSourceType.BEANLY_SPREADSHEET
    return ImportSourceType.GENERIC_SPREADSHEET


def _sheet(sheets: dict[str, list[dict[str, object]]], name: str) -> list[dict[str, object]]:
    return next((rows for title, rows in sheets.items() if title.casefold() == name.casefold()), [])


def _header(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _text(row: dict[str, object], key: str, default: str | None = None) -> str:
    value = str(row.get(key) or default or "").strip()
    if not value:
        raise ImportParseFailed(f"Required column is empty: {key}")
    return value


def _optional(value: object) -> str | None:
    normalized = str(value or "").strip()
    return normalized or None


def _strict_bool(value: object, *, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized in {"true", "1", "yes"}:
            return True
        if normalized in {"false", "0", "no"}:
            return False
    raise ImportParseFailed("Boolean fields must use true/false, yes/no, or 1/0")


def _slug(value: str) -> str:
    result = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    return result or hashlib.sha256(value.encode()).hexdigest()[:16]


def _major_to_minor(value: object) -> str:
    try:
        amount = Decimal(str(value)) * 100
    except (InvalidOperation, ValueError) as exc:
        raise ImportParseFailed("Price must be a decimal KZT amount") from exc
    if (
        not amount.is_finite()
        or amount != amount.to_integral_value()
        or not 0 <= amount <= MAX_NUMERIC_20_6_MINOR
    ):
        raise ImportParseFailed("Price is outside the supported exact minor-unit range")
    return str(int(amount))


def _decimal_text(value: object) -> str:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ImportParseFailed("Quantity must be decimal") from exc
    if not amount.is_finite() or amount <= 0:
        raise ImportParseFailed("Quantity must be finite and positive")
    return str(int(amount)) if amount == amount.to_integral_value() else format(amount, "f")


def _base_unit(unit: str) -> tuple[str, Decimal]:
    values = {
        "g": ("g", Decimal(1)),
        "kg": ("g", Decimal(1000)),
        "ml": ("ml", Decimal(1)),
        "l": ("ml", Decimal(1000)),
        "pcs": ("pcs", Decimal(1)),
        "pc": ("pcs", Decimal(1)),
    }
    try:
        return values[unit.strip().casefold()]
    except KeyError as exc:
        raise ImportParseFailed("Unit must be g, kg, ml, l or pcs") from exc
