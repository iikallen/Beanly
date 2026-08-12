import asyncio
import os
from collections.abc import AsyncIterator
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import pytest
import pytest_asyncio
from alembic import command
from alembic.config import Config
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.engine import make_url
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from beanly.core.database.session import get_session
from beanly.core.events.handlers.registry import EventHandlerRegistry
from beanly.core.events.outbox.dispatcher import OutboxDispatcher
from beanly.core.events.outbox.repositories import OutboxRepository
from beanly.core.security.audit import SecurityAuditEventModel, SecurityAuditRecorder
from beanly.main import app
from beanly.modules.analytics.application.projection_service import AnalyticsProjectionService
from beanly.modules.analytics.infrastructure.db.models import AnalyticsSalesDailyModel
from beanly.modules.analytics.infrastructure.db.repositories import SqlAlchemyAnalyticsRepository
from beanly.modules.analytics.infrastructure.handlers import register_analytics_handlers
from beanly.modules.analytics.infrastructure.source_reader import SqlAlchemyAnalyticsSourceReader
from beanly.modules.finance.application.projection_service import FinanceProjectionService
from beanly.modules.finance.infrastructure.db.models import FinanceEntryModel
from beanly.modules.finance.infrastructure.db.repositories import SqlAlchemyFinanceRepository
from beanly.modules.finance.infrastructure.handlers import register_finance_handlers
from beanly.modules.finance.infrastructure.source_reader import SqlAlchemyFinanceSourceReader
from beanly.modules.inventory.infrastructure.db.models import (
    InventoryItemModel,
    InventoryTransactionLineModel,
    InventoryTransactionModel,
    StockBalanceModel,
    WarehouseModel,
)
from beanly.modules.menu.infrastructure.db.models import (
    MenuCategoryModel,
    ModifierGroupModel,
    ModifierOptionModel,
    ProductModel,
    ProductVariantModel,
    RecipeComponentModel,
    RecipeModel,
)
from beanly.modules.onboarding.application.onboarding_service import OnboardingService
from beanly.modules.onboarding.infrastructure.db.models import (
    OnboardingImportRunModel,
    OnboardingStateModel,
)
from beanly.modules.onboarding.infrastructure.db.repositories import (
    SqlAlchemyOnboardingRepository,
)
from beanly.modules.onboarding.infrastructure.gateway import SqlAlchemyOnboardingGateway
from beanly.modules.onboarding.infrastructure.handlers import register_onboarding_handlers
from beanly.modules.payments.infrastructure.db.models import PaymentModel
from beanly.modules.sales.infrastructure.db.models import (
    PosRegisterModel,
    SalesOrderModel,
)


def _config(database_url: str) -> Config:
    config = Config(str(Path("alembic.ini")))
    config.set_main_option("sqlalchemy.url", database_url)
    return config


@pytest_asyncio.fixture
async def stage23_postgres_client() -> AsyncIterator[
    tuple[AsyncClient, async_sessionmaker[AsyncSession]]
]:
    source_url = os.getenv("POSTGRES_TEST_URL")
    if not source_url:
        pytest.skip("POSTGRES_TEST_URL is required for the PostgreSQL Stage 23 gate")
    source = make_url(source_url)
    database_name = f"beanly_stage23_{uuid4().hex}"
    admin_url = source.set(database="postgres").render_as_string(hide_password=False)
    database_url = source.set(database=database_name).render_as_string(hide_password=False)
    admin_engine = create_async_engine(admin_url, isolation_level="AUTOCOMMIT")
    engine = create_async_engine(database_url)
    sessions = async_sessionmaker(engine, expire_on_commit=False)
    try:
        async with admin_engine.connect() as connection:
            await connection.exec_driver_sql(f'CREATE DATABASE "{database_name}"')
        await asyncio.to_thread(command.upgrade, _config(database_url), "head")

        async def override_session() -> AsyncIterator[AsyncSession]:
            async with sessions() as session:
                yield session

        app.dependency_overrides[get_session] = override_session
        async with AsyncClient(
            transport=ASGITransport(app=app, raise_app_exceptions=False),
            base_url="http://test",
            headers={"origin": "http://localhost:3000"},
        ) as client:
            yield client, sessions
    finally:
        app.dependency_overrides.clear()
        await engine.dispose()
        async with admin_engine.connect() as connection:
            await connection.exec_driver_sql(
                f'DROP DATABASE IF EXISTS "{database_name}" WITH (FORCE)'
            )
        await admin_engine.dispose()


async def _workspace(client: AsyncClient, email: str):
    password = "correct-horse-battery-staple"
    assert (
        await client.post(
            "/api/v1/auth/register",
            json={
                "email": email,
                "password": password,
                "first_name": "Postgres",
                "last_name": "Owner",
            },
        )
    ).status_code == 201
    login = await client.post(
        "/api/v1/auth/login", json={"email": email, "password": password}
    )
    auth = {"authorization": f"Bearer {login.json()['access_token']}"}
    created = await client.post(
        "/api/v1/organizations",
        headers=auth,
        json={
            "name": "Postgres Coffee",
            "country_code": "KZ",
            "currency_code": "KZT",
            "first_location": {"name": "Dostyk", "timezone": "Asia/Almaty"},
        },
    )
    assert created.status_code == 201, created.text
    organization_id = UUID(created.json()["organization"]["id"])
    location_id = UUID(created.json()["location"]["id"])
    return {**auth, "X-Organization-ID": str(organization_id)}, organization_id, location_id


def _workbook(
    *,
    inventory_unit: str | None = None,
    opening_quantity=None,
    unit_cost_kzt=500,
    recipe_quantity=None,
    recipe_unit: str | None = None,
) -> bytes:
    workbook = Workbook()
    products = workbook.active
    products.title = "Products"
    products.append(["Category", "Product", "Variant", "SKU", "Price"])
    products.append(["Coffee", "Latte", "350", "LATTE-350", 1700])
    if inventory_unit:
        inventory = workbook.create_sheet("Inventory")
        inventory.append(["Name", "SKU", "Unit", "Opening Quantity", "Unit Cost KZT"])
        inventory.append(["Milk", "MILK", inventory_unit, opening_quantity, unit_cost_kzt])
    if recipe_quantity is not None and recipe_unit is not None:
        recipes = workbook.create_sheet("Recipes")
        recipes.append(["Product", "Variant", "Inventory Item", "Quantity", "Unit"])
        recipes.append(["Latte", "350", "Milk", recipe_quantity, recipe_unit])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _upload(
    client: AsyncClient,
    headers: dict[str, str],
    location_id: UUID,
    content: bytes,
) -> dict:
    response = await client.post(
        "/api/v1/onboarding/imports",
        headers=headers,
        data={
            "client_import_id": str(uuid4()),
            "location_id": str(location_id),
            "source_type": "BEANLY_SPREADSHEET",
        },
        files={
            "file": (
                "beanly-menu-template.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


@pytest.mark.anyio
async def test_postgres_bootstrap_and_apply_are_concurrently_idempotent(
    stage23_postgres_client,
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-concurrency@example.com"
    )
    bootstraps = await asyncio.gather(
        client.post("/api/v1/onboarding/bootstrap", headers=headers, json={}),
        client.post("/api/v1/onboarding/bootstrap", headers=headers, json={}),
    )
    assert [value.status_code for value in bootstraps] == [200, 200]
    assert len({value.json()["warehouse_id"] for value in bootstraps}) == 1
    assert len({value.json()["register_id"] for value in bootstraps}) == 1
    async with sessions() as session:
        assert await session.scalar(
            select(func.count()).select_from(WarehouseModel).where(
                WarehouseModel.organization_id == organization_id
            )
        ) == 1
        assert await session.scalar(
            select(func.count()).select_from(PosRegisterModel).where(
                PosRegisterModel.organization_id == organization_id
            )
        ) == 1

    run = await _upload(client, headers, location_id, _workbook())
    assert run["status"] == "READY"
    applies = await asyncio.gather(
        client.post(f"/api/v1/onboarding/imports/{run['id']}/apply", headers=headers),
        client.post(f"/api/v1/onboarding/imports/{run['id']}/apply", headers=headers),
    )
    assert [value.status_code for value in applies] == [200, 200]
    assert {value.json()["status"] for value in applies} == {"APPLIED"}
    async with sessions() as session:
        assert await session.scalar(
            select(func.count()).select_from(ProductModel).where(
                ProductModel.organization_id == organization_id
            )
        ) == 1


@pytest.mark.anyio
async def test_postgres_unit_conflict_rolls_back_the_entire_apply(
    stage23_postgres_client,
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-rollback@example.com"
    )
    await client.post("/api/v1/onboarding/bootstrap", headers=headers, json={})
    existing = await client.post(
        "/api/v1/inventory/items",
        headers=headers,
        # The name intentionally differs: exact SKU alone must never override a unit conflict.
        json={"name": "Whole Milk", "sku": "MILK", "base_unit": "pcs"},
    )
    assert existing.status_code == 201, existing.text
    run = await _upload(
        client,
        headers,
        location_id,
        _workbook(inventory_unit="ml"),
    )
    response = await client.post(
        f"/api/v1/onboarding/imports/{run['id']}/apply", headers=headers
    )
    assert response.status_code in {409, 422}
    async with sessions() as session:
        assert await session.scalar(
            select(func.count()).select_from(MenuCategoryModel).where(
                MenuCategoryModel.organization_id == organization_id
            )
        ) == 0
        assert await session.scalar(
            select(func.count()).select_from(ProductModel).where(
                ProductModel.organization_id == organization_id
            )
        ) == 0
        assert await session.scalar(
            select(func.count()).select_from(InventoryItemModel).where(
                InventoryItemModel.organization_id == organization_id
            )
        ) == 1
        status = await session.scalar(
            select(OnboardingImportRunModel.status).where(
                OnboardingImportRunModel.id == UUID(run["id"])
            )
        )
    assert status == "FAILED"


@pytest.mark.anyio
async def test_postgres_opening_balance_uses_ledger_and_projection(
    stage23_postgres_client,
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-opening@example.com"
    )
    await client.post("/api/v1/onboarding/bootstrap", headers=headers, json={})
    run = await _upload(
        client,
        headers,
        location_id,
        _workbook(
            inventory_unit="kg",
            opening_quantity=8.4,
            unit_cost_kzt=8000,
            recipe_quantity=0.018,
            recipe_unit="kg",
        ),
    )
    response = await client.post(
        f"/api/v1/onboarding/imports/{run['id']}/apply", headers=headers
    )
    assert response.status_code == 200, response.text
    async with sessions() as session:
        transaction = await session.scalar(
            select(InventoryTransactionModel).where(
                InventoryTransactionModel.organization_id == organization_id,
                InventoryTransactionModel.type == "OPENING_BALANCE",
            )
        )
        assert transaction is not None and transaction.status == "POSTED"
        line = (
            await session.execute(
                select(
                    InventoryTransactionLineModel.quantity_delta,
                    InventoryTransactionLineModel.unit_cost_amount,
                    InventoryTransactionLineModel.total_cost_amount,
                ).where(
                InventoryTransactionLineModel.transaction_id == transaction.id
            )
            )
        ).one()
        projected = await session.scalar(
            select(StockBalanceModel.quantity).where(
                StockBalanceModel.organization_id == organization_id
            )
        )
        recipe_quantity = await session.scalar(select(RecipeComponentModel.quantity))
    assert str(line.quantity_delta) == str(projected) == "8400.000000"
    assert str(line.unit_cost_amount) == "8.000000"
    assert str(line.total_cost_amount) == "67200.000000"
    assert str(recipe_quantity) == "18.000000"


@pytest.mark.anyio
async def test_postgres_inventory_only_workbook_previews_and_applies_without_menu_facts(
    stage23_postgres_client,
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-inventory-only@example.com"
    )
    await client.post("/api/v1/onboarding/bootstrap", headers=headers, json={})
    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "Inventory"
    inventory.append(["Name", "SKU", "Unit", "Opening Quantity", "Unit Cost KZT"])
    inventory.append(["Coffee Beans", "BEANS", "kg", 10, 8000])
    buffer = BytesIO()
    workbook.save(buffer)
    run = await _upload(client, headers, location_id, buffer.getvalue())
    assert run["status"] == "READY"
    response = await client.post(
        f"/api/v1/onboarding/imports/{run['id']}/apply", headers=headers
    )
    assert response.status_code == 200, response.text
    async with sessions() as session:
        assert await session.scalar(
            select(func.count()).select_from(ProductModel).where(
                ProductModel.organization_id == organization_id
            )
        ) == 0
        assert await session.scalar(
            select(func.count()).select_from(InventoryItemModel).where(
                InventoryItemModel.organization_id == organization_id
            )
        ) == 1
        quantity = await session.scalar(
            select(StockBalanceModel.quantity).where(
                StockBalanceModel.organization_id == organization_id
            )
        )
    assert str(quantity) == "10000.000000"


@pytest.mark.anyio
async def test_postgres_classic_template_preview_validates_and_applies_draft_facts(
    stage23_postgres_client,
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-classic@example.com"
    )
    await client.post("/api/v1/onboarding/bootstrap", headers=headers, json={})
    preview = await client.post(
        "/api/v1/onboarding/templates/classic_coffee_shop/preview",
        headers=headers,
        json={
            "client_import_id": str(uuid4()),
            "version": 1,
            "location_id": str(location_id),
            "options": {
                "sizes": ["250", "350", "450"],
                "alternative_milks": ["Oat milk"],
                "extras": ["Extra shot"],
                "packaging": True,
                "include_draft_recipes": True,
            },
        },
    )
    assert preview.status_code == 200, preview.text
    run_id = preview.json()["id"]
    validated = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/validate", headers=headers
    )
    assert validated.status_code == 200, validated.text
    assert validated.json()["valid"] is True
    applied = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/apply", headers=headers
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["status"] == "APPLIED"
    async with sessions() as session:
        provenance = (
            await session.execute(
                select(
                    OnboardingImportRunModel.source_name,
                    OnboardingImportRunModel.source_version,
                ).where(OnboardingImportRunModel.id == UUID(run_id))
            )
        ).one()
        assert provenance == ("classic_coffee_shop", 1)
        assert await session.scalar(
            select(func.count()).select_from(MenuCategoryModel).where(
                MenuCategoryModel.organization_id == organization_id
            )
        ) == 4
        products = list(
            await session.scalars(
                select(ProductModel).where(ProductModel.organization_id == organization_id)
            )
        )
        variants = list(
            await session.scalars(
                select(ProductVariantModel).where(
                    ProductVariantModel.organization_id == organization_id
                )
            )
        )
        assert len(products) >= 10 and {value.status for value in products} == {"DRAFT"}
        assert len(variants) >= len(products) and {value.status for value in variants} == {
            "DRAFT"
        }
        assert await session.scalar(
            select(func.count()).select_from(InventoryItemModel).where(
                InventoryItemModel.organization_id == organization_id
            )
        ) >= 4
        assert await session.scalar(select(func.count()).select_from(RecipeModel)) >= 1
        assert await session.scalar(select(func.count()).select_from(ModifierGroupModel)) >= 1
        assert await session.scalar(select(func.count()).select_from(ModifierOptionModel)) >= 1


@pytest.mark.anyio
async def test_postgres_new_workspace_reaches_first_sale_and_all_stage23_projections(
    stage23_postgres_client, monkeypatch
) -> None:
    client, sessions = stage23_postgres_client
    headers, organization_id, location_id = await _workspace(
        client, "stage23-first-sale@example.com"
    )
    bootstrap = await client.post("/api/v1/onboarding/bootstrap", headers=headers, json={})
    assert bootstrap.status_code == 200, bootstrap.text
    preview = await client.post(
        "/api/v1/onboarding/templates/classic_coffee_shop/preview",
        headers=headers,
        json={
            "client_import_id": str(uuid4()),
            "version": 1,
            "location_id": str(location_id),
            "options": {
                "sizes": ["350"],
                "alternative_milks": [],
                "extras": [],
                "packaging": False,
                "include_draft_recipes": True,
            },
        },
    )
    assert preview.status_code == 200, preview.text
    run_id = preview.json()["id"]
    validated = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/validate", headers=headers
    )
    assert validated.status_code == 200 and validated.json()["valid"] is True
    applied = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/apply", headers=headers
    )
    assert applied.status_code == 200 and applied.json()["status"] == "APPLIED"

    async with sessions() as session:
        cappuccino = await session.scalar(
            select(ProductModel).where(
                ProductModel.organization_id == organization_id,
                ProductModel.name == "Cappuccino",
            )
        )
        assert cappuccino is not None
        variant = await session.scalar(
            select(ProductVariantModel).where(ProductVariantModel.product_id == cappuccino.id)
        )
        assert variant is not None
    not_confirmed = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/activate-ready",
        headers=headers,
        json={
            "product_ids": [str(cappuccino.id)],
            "confirm_starter_recipes_reviewed": False,
        },
    )
    assert not_confirmed.status_code == 422, not_confirmed.text
    activated = await client.post(
        f"/api/v1/onboarding/imports/{run_id}/activate-ready",
        headers=headers,
        json={
            "product_ids": [str(cappuccino.id)],
            "confirm_starter_recipes_reviewed": True,
        },
    )
    assert activated.status_code == 200, activated.text
    assert activated.json()["activated_count"] == 1

    shift = await client.post(
        "/api/v1/sales/shifts/open",
        headers=headers,
        json={
            "register_id": bootstrap.json()["register_id"],
            "warehouse_id": bootstrap.json()["warehouse_id"],
        },
    )
    assert shift.status_code == 201, shift.text
    order = await client.post(
        "/api/v1/sales/orders",
        headers=headers,
        json={
            "client_order_id": str(uuid4()),
            "shift_id": shift.json()["id"],
            "order_type": "TAKEAWAY",
        },
    )
    assert order.status_code == 201, order.text
    order = await client.post(
        f"/api/v1/sales/orders/{order.json()['id']}/items",
        headers=headers,
        json={
            "client_item_id": str(uuid4()),
            "variant_id": str(variant.id),
            "selected_option_ids": [],
            "quantity": 1,
        },
    )
    assert order.status_code == 201, order.text
    amount_minor = order.json()["total_minor"]
    payment = await client.post(
        f"/api/v1/payments/orders/{order.json()['id']}/complete",
        headers=headers,
        json={
            "client_payment_id": str(uuid4()),
            "lines": [
                {
                    "method": "CASH",
                    "amount_minor": amount_minor,
                    "cash_received_minor": amount_minor,
                }
            ],
        },
    )
    assert payment.status_code == 201, payment.text

    class Histogram:
        def __init__(self) -> None:
            self.values: list[float] = []

        def record(self, value: float) -> None:
            self.values.append(value)

    first_sale = Histogram()
    monkeypatch.setattr(
        "beanly.modules.onboarding.application.onboarding_service.metrics.onboarding_time_to_first_sale",
        first_sale,
    )
    async with sessions() as session:
        registry = EventHandlerRegistry()
        register_finance_handlers(
            registry,
            FinanceProjectionService(
                SqlAlchemyFinanceRepository(session), SqlAlchemyFinanceSourceReader(session)
            ),
        )
        register_analytics_handlers(
            registry,
            AnalyticsProjectionService(
                SqlAlchemyAnalyticsRepository(session),
                SqlAlchemyAnalyticsSourceReader(session),
            ),
        )
        register_onboarding_handlers(
            registry,
            OnboardingService(
                SqlAlchemyOnboardingRepository(session),
                SqlAlchemyOnboardingGateway(
                    session, live_transport_enabled=False, nkt_configured=False
                ),
            ),
            SecurityAuditRecorder(session),
        )
        dispatcher = OutboxDispatcher(
            OutboxRepository(session), registry, "stage23-first-sale", batch_size=100
        )
        while await dispatcher.run_once():
            pass

    async with sessions() as session:
        persisted_order = await session.get(SalesOrderModel, UUID(order.json()["id"]))
        persisted_payment = await session.get(PaymentModel, UUID(payment.json()["id"]))
        onboarding = await session.scalar(
            select(OnboardingStateModel).where(
                OnboardingStateModel.organization_id == organization_id
            )
        )
        assert persisted_order is not None and persisted_order.status == "PAID"
        assert persisted_payment is not None
        assert await session.scalar(
            select(func.count()).select_from(InventoryTransactionModel).where(
                InventoryTransactionModel.organization_id == organization_id,
                InventoryTransactionModel.type == "SALE",
            )
        ) == 1
        assert await session.scalar(
            select(func.count()).select_from(FinanceEntryModel).where(
                FinanceEntryModel.organization_id == organization_id
            )
        ) >= 1
        analytics = await session.scalar(
            select(AnalyticsSalesDailyModel).where(
                AnalyticsSalesDailyModel.organization_id == organization_id
            )
        )
        assert analytics is not None and analytics.paid_orders == 1
        assert onboarding is not None and onboarding.status == "COMPLETED"
        assert await session.scalar(
            select(func.count()).select_from(SecurityAuditEventModel).where(
                SecurityAuditEventModel.organization_id == organization_id,
                SecurityAuditEventModel.action == "ONBOARDING_COMPLETED",
            )
        ) == 1
    assert len(first_sale.values) == 1
