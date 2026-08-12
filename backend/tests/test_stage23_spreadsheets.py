from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from beanly.modules.onboarding.application.template_service import TemplateService
from beanly.modules.onboarding.domain.enums import (
    ImportEntityType,
    ImportSourceType,
    UploadSourceType,
)
from beanly.modules.onboarding.domain.exceptions import ImportParseFailed
from beanly.modules.onboarding.infrastructure.spreadsheets import (
    inspect_spreadsheet,
    parse_spreadsheet,
)


def _workbook(headers: list[str], rows: list[list[object]], title: str = "Products") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _beanly_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    values = {
        "Products": (
            [
                "Category",
                "Product",
                "Variant",
                "SKU",
                "Price",
                "Location",
                "Available",
                "Description",
            ],
            [["Coffee", "Latte", "350", "LATTE-350", 1700, "Main", True, "Milk coffee"]],
        ),
        "Inventory": (
            ["Name", "SKU", "Unit", "Opening Quantity", "Unit Cost KZT"],
            [["Coffee Beans", "BEANS", "kg", 8.4, 8000]],
        ),
        "Recipes": (
            ["Product", "Variant", "Inventory Item", "Quantity", "Unit"],
            [["Latte", "350", "Coffee Beans", 18, "g"]],
        ),
        "Modifiers": (
            [
                "Product",
                "Variant",
                "Group",
                "Selection Type",
                "Option",
                "Price Delta",
                "Inventory Item",
                "Quantity Delta",
                "Unit",
            ],
            [["Latte", "350", "Extras", "MULTIPLE", "Extra shot", 400, "Coffee Beans", 18, "g"]],
        ),
    }
    for title, (headers, rows) in values.items():
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_documented_poster_columns_create_unique_product_modifications() -> None:
    content = _workbook(
        [
            "Name",
            "Category",
            "Station",
            "Type",
            "Price",
            "Weight item",
            "PosterID product_id",
            "PosterID modificator_id",
        ],
        [
            ["Cappuccino 350", "Coffee", "Bar", "Dish", 1700, 350, 512, 1],
            ["Cappuccino 450", "Coffee", "Bar", "Dish", 1900, 450, 512, 2],
        ],
    )
    draft, source_type, _ = parse_spreadsheet(
        content,
        "poster-export.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.POSTER,
    )
    assert source_type is ImportSourceType.POSTER_EXPORT
    products = [
        entity for entity in draft.entities if entity.entity_type is ImportEntityType.PRODUCT
    ]
    variants = [
        entity for entity in draft.entities if entity.entity_type is ImportEntityType.VARIANT
    ]
    assert len(products) == 1
    assert len(variants) == 2
    assert len({entity.source_key for entity in draft.entities}) == len(draft.entities)
    assert products[0].payload["source_external_id"] == "512"
    assert {value.payload["source_external_id"] for value in variants} == {"1", "2"}
    assert all("POSTER_REAL_FIXTURE_UNVERIFIED" in value.warning_codes for value in products)


def test_official_poster_russian_headers_and_recipe_composition_are_canonicalized() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tech cards"
    sheet.append([None, None, None, None, None])
    sheet.append(
        [
            "PosterID product_id (не менять!)",
            "Название блюда",
            "Состав",
            "Брутто, г",
            "Цена",
        ]
    )
    sheet.append([112, "Цезарь", "Курица", "150,5", 2400])
    buffer = BytesIO()
    workbook.save(buffer)

    draft, source_type, _ = parse_spreadsheet(
        buffer.getvalue(),
        "poster-recipes.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.POSTER,
    )

    assert source_type is ImportSourceType.POSTER_EXPORT
    inventory = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.INVENTORY_ITEM
    )
    recipe = next(
        entity for entity in draft.entities if entity.entity_type is ImportEntityType.RECIPE
    )
    assert inventory.payload == {"name": "Курица", "sku": None, "base_unit": "g"}
    assert recipe.payload["components"] == [
        {
            "inventory_item_key": inventory.source_key,
            "quantity": "150.5",
            "unit": "g",
        }
    ]


def test_official_poster_ingredient_headers_create_opening_balance_in_base_units() -> None:
    content = _workbook(
        ["Name", "Category", "Unit", "Inventory count", "Value"],
        [["Squid", "Seafood", "kg", "1.300 kg", 18]],
        title="Ingredients",
    )
    draft, _, _ = parse_spreadsheet(
        content,
        "poster-ingredients.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.POSTER,
    )
    inventory = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.INVENTORY_ITEM
    )
    opening = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.OPENING_BALANCE
    )
    assert inventory.payload["base_unit"] == "g"
    assert opening.payload["quantity"] == "1300"
    assert opening.payload["unit_cost_minor"] == "1800"


@pytest.mark.parametrize("value", ("=2+2", "+cmd", "-1+2", "@SUM(A1:A2)"))
def test_formula_like_spreadsheet_text_is_rejected(value: str) -> None:
    content = _workbook(
        ["Category", "Product", "Variant", "Price"],
        [["Coffee", value, "350", 1700]],
    )
    with pytest.raises(ImportParseFailed):
        parse_spreadsheet(
            content,
            "beanly.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            UploadSourceType.BEANLY_SPREADSHEET,
        )


def test_duplicate_headers_and_broken_utf8_are_rejected() -> None:
    duplicate_headers = _workbook(
        ["Category", "Product", "Product", "Price"],
        [["Coffee", "Latte", "350", 1700]],
    )
    with pytest.raises(ImportParseFailed):
        parse_spreadsheet(
            duplicate_headers,
            "menu.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            UploadSourceType.AUTO,
        )
    with pytest.raises(ImportParseFailed):
        parse_spreadsheet(
            b"Category,Product,Price\nCoffee,Latte,\xff",
            "menu.csv",
            "text/csv",
            UploadSourceType.AUTO,
        )


def test_official_workbook_headers_round_trip_into_every_canonical_section() -> None:
    generated = load_workbook(BytesIO(TemplateService().workbook()), read_only=True)
    assert tuple(generated.sheetnames) == ("Products", "Recipes", "Inventory", "Modifiers")
    assert [cell.value for cell in generated["Products"][1]] == [
        "Category",
        "Product",
        "Variant",
        "SKU",
        "Price",
        "Location",
        "Available",
        "Description",
    ]
    assert [cell.value for cell in generated["Inventory"][1]] == [
        "Name",
        "SKU",
        "Unit",
        "Opening Quantity",
        "Unit Cost KZT",
    ]

    draft, source_type, _ = parse_spreadsheet(
        _beanly_workbook(),
        "beanly-menu-template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.BEANLY_SPREADSHEET,
    )
    assert source_type is ImportSourceType.BEANLY_SPREADSHEET
    types = {entity.entity_type for entity in draft.entities}
    assert {
        ImportEntityType.CATEGORY,
        ImportEntityType.PRODUCT,
        ImportEntityType.VARIANT,
        ImportEntityType.INVENTORY_ITEM,
        ImportEntityType.RECIPE,
        ImportEntityType.MODIFIER_GROUP,
        ImportEntityType.MODIFIER_OPTION,
        ImportEntityType.LOCATION_PRICE,
        ImportEntityType.OPENING_BALANCE,
    } <= types


def test_opening_balance_converts_kg_to_base_grams_without_direct_stock_fact() -> None:
    draft, _, _ = parse_spreadsheet(
        _beanly_workbook(),
        "beanly-menu-template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.BEANLY_SPREADSHEET,
    )
    item = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.INVENTORY_ITEM
    )
    opening = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.OPENING_BALANCE
    )
    assert item.payload["base_unit"] == "g"
    assert opening.payload["quantity"] == "8400"
    assert opening.payload["unit"] == "g"
    assert opening.payload["unit_cost_minor"] == "800000"


def test_available_false_text_is_not_coerced_to_true() -> None:
    content = _workbook(
        ["Category", "Product", "Variant", "Price", "Location", "Available"],
        [["Coffee", "Latte", "350", 1700, "Main", "false"]],
    )

    draft, _, _ = parse_spreadsheet(
        content,
        "beanly.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.BEANLY_SPREADSHEET,
    )
    location_price = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.LOCATION_PRICE
    )

    assert location_price.payload["available"] is False


def test_available_rejects_ambiguous_boolean_text() -> None:
    content = _workbook(
        ["Category", "Product", "Variant", "Price", "Location", "Available"],
        [["Coffee", "Latte", "350", 1700, "Main", "perhaps"]],
    )

    with pytest.raises(ImportParseFailed, match="Boolean fields"):
        parse_spreadsheet(
            content,
            "beanly.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            UploadSourceType.BEANLY_SPREADSHEET,
        )


@pytest.mark.parametrize(
    "mapping",
    (
        {
            "Наименование": "product",
            "Дубликат": "product",
            "Группа": "category",
            "Цена": "price",
        },
        {"Нет такого столбца": "product"},
        {"Наименование": "unsupported_business_field"},
    ),
)
def test_generic_mapping_rejects_ambiguous_or_unknown_columns(
    mapping: dict[str, str],
) -> None:
    content = "Наименование,Дубликат,Группа,Цена\nLatte,Latte 2,Coffee,1700\n".encode()
    with pytest.raises(ImportParseFailed):
        parse_spreadsheet(
            content,
            "generic.csv",
            "text/csv",
            UploadSourceType.GENERIC_SPREADSHEET,
            mapping,
        )


def test_inspect_does_not_misclassify_arbitrary_products_sheet_as_beanly() -> None:
    content = _workbook(
        ["Наименование", "Группа", "Цена"],
        [["Latte", "Coffee", 1700]],
    )
    result = inspect_spreadsheet(
        content,
        "generic.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.AUTO,
    )
    assert result["source_type"] is ImportSourceType.GENERIC_SPREADSHEET
    assert result["mapping_required"] is True


def test_inventory_only_official_workbook_uses_the_shared_canonical_pipeline() -> None:
    content = _workbook(
        ["Name", "SKU", "Unit", "Opening Quantity", "Unit Cost KZT"],
        [["Coffee Beans", "BEANS", "kg", 10, 8000]],
        title="Inventory",
    )
    draft, source_type, _ = parse_spreadsheet(
        content,
        "beanly-inventory.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        UploadSourceType.BEANLY_SPREADSHEET,
    )
    assert source_type is ImportSourceType.BEANLY_SPREADSHEET
    assert {entity.entity_type for entity in draft.entities} == {
        ImportEntityType.INVENTORY_ITEM,
        ImportEntityType.OPENING_BALANCE,
    }
    opening = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.OPENING_BALANCE
    )
    assert opening.payload["quantity"] == "10000"
    assert opening.payload["unit"] == "g"


def test_generic_inventory_csv_mapping_does_not_require_fake_product_fields() -> None:
    content = "Товар,Ед.,Остаток,Себестоимость\nМолоко,L,2.5,750\n".encode()
    draft, source_type, _ = parse_spreadsheet(
        content,
        "inventory.csv",
        "text/csv",
        UploadSourceType.GENERIC_SPREADSHEET,
        {
            "Товар": "name",
            "Ед.": "unit",
            "Остаток": "opening quantity",
            "Себестоимость": "unit cost kzt",
        },
    )
    assert source_type is ImportSourceType.GENERIC_SPREADSHEET
    assert {entity.entity_type for entity in draft.entities} == {
        ImportEntityType.INVENTORY_ITEM,
        ImportEntityType.OPENING_BALANCE,
    }
    item = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.INVENTORY_ITEM
    )
    opening = next(
        entity
        for entity in draft.entities
        if entity.entity_type is ImportEntityType.OPENING_BALANCE
    )
    assert item.payload["base_unit"] == "ml"
    assert opening.payload["quantity"] == "2500"
